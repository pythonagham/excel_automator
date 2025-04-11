import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os

def automate_attendance():
    try:
        # Step 1: Load the workbook and sheet
        wb = openpyxl.load_workbook(excel_path.get())
        sheet = wb.active

        # Step 2: Read attendees' names
        with open(txt_path.get(), 'r') as file:
            attendees = [line.strip().lower() for line in file.readlines()]

        # Step 3: Find the first empty column
        empty_column = None
        for col in range(1, sheet.max_column + 2):
            if all(sheet.cell(row=row, column=col).value is None for row in range(2, sheet.max_row + 1)):
                empty_column = col
                break

        if not empty_column:
            messagebox.showerror("Error", "No empty column found for the session.")
            return

        # Step 4: Mark presence
        for row in range(2, sheet.max_row + 1):
            first = sheet.cell(row=row, column=1).value
            last = sheet.cell(row=row, column=2).value
            if first and last:
                full_name = f"{first} {last}".strip().lower()
                if full_name in attendees:
                    sheet.cell(row=row, column=empty_column).value = 'P'

        # Step 5: Locate Total and % Present cells
        total_row = percent_row = None
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                val = sheet.cell(row=row, column=col).value
                if val == "Total":
                    total_row = row
                elif val == "Percentage":
                    percent_row = row

        last_student_row = total_row - 1 if total_row else sheet.max_row
        col_letter = openpyxl.utils.get_column_letter(empty_column)

        # Step 6: Add formulas
        if total_row:
            sheet.cell(row=total_row, column=empty_column).value = f'=COUNTIF({col_letter}2:{col_letter}{last_student_row}, "P")'
        if percent_row:
            sheet.cell(row=percent_row, column=empty_column).value = f'=ROUND({col_letter}{total_row}/COUNTA(A2:A{last_student_row})*100,2)'

        # Step 7: Save workbook
        base, ext = os.path.splitext(excel_path.get())
        output_path = base + "_updated" + ext
        wb.save(output_path)

        messagebox.showinfo("Success", f"Attendance recorded and saved to:\n{output_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))

def browse_txt():
    path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if path:
        txt_path.set(path)

def browse_excel():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if path:
        excel_path.set(path)

# GUI setup
root = tk.Tk()
root.title("Attendance Automator")
root.geometry("700x450")
root.resizable(False, False)
root.configure(bg='#acffa2')

#Logo
logo =tk.PhotoImage(file="logo.png")
tk.Label(root,image=logo,bg="#acffa2").place(x=300,y=20)

#heading
heading=tk.Label(root,text="Excel Automator",
              font='arial 20 bold',fg="#03723d",
              bg="#acffa2")
heading.place(x=240,y=135)


txt_path = tk.StringVar()
excel_path = tk.StringVar()

tk.Label(root, text="Select Attendance TXT File:",font='arial 11 bold',fg="#03723d",bg="#acffa2").place(x=43, y=204)
tk.Entry(root, textvariable=txt_path, width=30,font='arial 12 bold').place(x=250, y=204)
tk.Button(root, text="Browse", width=9,
            cursor='hand2', bg="#03ad2c",fg='white', bd=0,
          activebackground='#a2d604',
            font='arial 11 bold',command=browse_txt).place(x=530, y=204)

tk.Label(root, text="Select Students Excel File:",font='arial 11 bold',fg="#03723d",bg="#acffa2").place(x=43, y=264)
tk.Entry(root, textvariable=excel_path, width=30,font='arial 12 bold').place(x=250, y=264)
tk.Button(root, text="Browse",width=9,
            cursor='hand2', bg="#03ad2c",fg='white', bd=0,
            activebackground='#a2d604',
            font='arial 11 bold', command=browse_excel).place(x=530, y=264)

tk.Button(root, text="Automate", width=12,height=2,
            cursor='hand2', bg="#03723d",fg='white', bd=0,
            activebackground='#a2d604',
            font='arial 13 bold',command=automate_attendance).place(x=285, y=324)

insta_page=tk.Label(root,text="@pythonagham",bg='#acffa2',
              fg='black',font='arial 10 bold italic')
insta_page.place(x=300,y=410)

root.mainloop()
